from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl


# Create your views here.
from Remote_User.models import review_Model,ClientRegister_Model,finding_psychology_model,finding_psychological_instability_model,recommend_Model,psychology_accuracy_model


def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:

            enter = ClientRegister_Model.objects.get(username=username, password=password)
            request.session["userid"] = enter.id
            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

            finding_psychological_instability_model.objects.all().delete()
            psychology_accuracy_model.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        finding_psychological_instability_model.objects.create(
        names=active_sheet.cell(r, 1).value,
        age=active_sheet.cell(r, 2).value,
        gender=active_sheet.cell(r, 3).value,
        country=active_sheet.cell(r, 4).value,
        state=active_sheet.cell(r, 5).value,
        self_employeed=active_sheet.cell(r, 6).value,
        family_history=active_sheet.cell(r, 7).value,
        treatment=active_sheet.cell(r, 8).value,
        work_interferes=active_sheet.cell(r, 9).value,
        No_of_employees=active_sheet.cell(r, 10).value,
        remote_network=active_sheet.cell(r, 10).value,
        technical_company=active_sheet.cell(r, 12).value,
        benefits=active_sheet.cell(r, 13).value,
        care_options=active_sheet.cell(r, 14).value,
        welness_program=active_sheet.cell(r, 15).value,
        seek_help=active_sheet.cell(r, 16).value,
        anonymity=active_sheet.cell(r, 17).value,
        leave1=active_sheet.cell(r, 18).value,
        mental_health_consequence=active_sheet.cell(r, 19).value,
        physical_health_consequence=active_sheet.cell(r, 20).value,
        co_workers=active_sheet.cell(r, 21).value,
        supervisor=active_sheet.cell(r, 22).value,
        mental_health_interview=active_sheet.cell(r, 23).value,
        physical_health_interview=active_sheet.cell(r, 24).value,
        mental_health=active_sheet.cell(r, 25).value,
        obs_consequence=active_sheet.cell(r, 26).value,
        Remarks_or_comments=active_sheet.cell(r, 27).value,
        dislikes=active_sheet.cell(r, 28).value,
        likes=active_sheet.cell(r, 29).value,
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:

        return render(request,'RUser/Register1.html')


def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Search_PsychologyDataSets_Details(request):

    if request.method == "POST":
        kword = request.POST.get('keyword')
        obj = finding_psychological_instability_model.objects.all().filter(names__contains=kword)
        return render(request, 'RUser/Search_PsychologyDataSets_Details.html',{'objs': obj})
    return render(request, 'RUser/Search_PsychologyDataSets_Details.html')

def ratings(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = finding_psychological_instability_model.objects.get(id=pk)
    unid = objs.id
    vot_count = finding_psychological_instability_model.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.ratings
        vott1 = vott + 1
        obj = get_object_or_404(finding_psychological_instability_model, id=unid)
        obj.ratings = vott1
        obj.save(update_fields=["ratings"])
        return redirect('Add_DataSet_Details')

    return render(request,'RUser/ratings.html',{'objs':vott1})



